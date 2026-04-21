"""FFPE training with +1 pseudocount normalization: x = log2((r+1) / gm_mean(hk+1))."""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl, numpy as np
from common import DATA_DIR, DAT_INTEG, GENE_NAMES, HK_NAMES

FFPE_XLSX = 'Previous_calibration_run/original_ffpe_frozen_paired/Results Study 20220706 UCSF UDR3097.xlsx'

wb = openpyxl.load_workbook(FFPE_XLSX, read_only=True, data_only=True)
ws = wb['RALEIGH_75_C9543']
rows = list(ws.iter_rows(values_only=True))
sample_ids = [x for x in rows[2][3:] if x is not None]
first_sample_col = 3
n_samples = len(sample_ids)
probe_to_row = {r[1]: i for i, r in enumerate(rows)
                if r[0] in ('Endogenous', 'Housekeeping') and r[1]}
X_raw = np.array([[float(rows[probe_to_row[g]][first_sample_col + j] or 0)
                   for g in GENE_NAMES] for j in range(n_samples)])
HK    = np.array([[float(rows[probe_to_row[h]][first_sample_col + j] or 0)
                   for h in HK_NAMES]   for j in range(n_samples)])

# +1 pseudocount normalization (same as common.hk_normalize)
h = np.exp(np.mean(np.log(HK + 1), axis=1))
X_ffpe = np.log2((X_raw + 1) / h[:, None])
print(f"X_ffpe (+1 pseudocount): shape {X_ffpe.shape}, mean {X_ffpe.mean():.3f}, sd {X_ffpe.std():.3f}")

wb2 = openpyxl.load_workbook(DAT_INTEG, read_only=True, data_only=True)
ws2 = wb2['Sheet1']
hdr = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
idx_id, idx_site = hdr.index('ID'), hdr.index('Site')
idx_dc, idx_chen = hdr.index('Data.Chen'), hdr.index('Chen')
id_to_chen = {str(row[idx_id]): row[idx_chen] for row in ws2.iter_rows(min_row=2, values_only=True)
              if row[idx_site] == 'UCSF' and row[idx_dc] == 1}
keep_idx, y_list, kept_ids = [], [], []
for i, sid in enumerate(sample_ids):
    if sid in id_to_chen and id_to_chen[sid] is not None:
        keep_idx.append(i); y_list.append(id_to_chen[sid]); kept_ids.append(sid)
X_train = X_ffpe[keep_idx]; y_train = np.array(y_list, dtype=float)
print(f"FFPE-paired training: N = {X_train.shape[0]}")

np.savez(os.path.join(DATA_DIR, 'ffpe_paired_training_pc.npz'),
         X_norm=X_train, y=y_train, sample_ids=np.array(kept_ids),
         gene_names=np.array(GENE_NAMES), hk_names=np.array(HK_NAMES))
csv_path = os.path.join(DATA_DIR, 'ffpe_paired_training_pc.csv')
with open(csv_path, 'w') as f:
    f.write('y,' + ','.join(GENE_NAMES) + '\n')
    for i in range(X_train.shape[0]):
        f.write(f'{y_train[i]},' + ','.join(f'{x:.6f}' for x in X_train[i]) + '\n')
print(f"Saved {csv_path}")
