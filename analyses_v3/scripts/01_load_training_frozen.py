"""Load 173-sample UCSF training cohort from dat.integ.master.

Biomarker cols 121-154 (34 genes; LINC02593 at col 143) are raw counts.
HK cols 155-161 are raw counts. Applies standard Chen 2023 HK normalization.

Output: analyses/data/training_173.npz with keys:
  X_norm (173x34), X_raw (173x34), HK (173x7), y (173,),
  sample_ids, gene_names, hk_names.
"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
import numpy as np
from common import (DAT_INTEG, DATA_DIR, GENE_NAMES, HK_NAMES,
                    hk_normalize, ensure_dirs)

ensure_dirs()
wb = openpyxl.load_workbook(DAT_INTEG, read_only=True, data_only=True)
ws = wb['Sheet1']
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

# Column indexing. Biomarker genes are at fixed positions; LINC02593 has two
# header cells ('LINC02593...82' and 'LINC02593...144'). The first (col 81) is a
# pre-computed log2 HK-normalized value; the second (col 143) is the raw count.
# Use col 143 for consistency: normalize ourselves from raw counts.
gene_cols = {}
for g in GENE_NAMES:
    if g == 'LINC02593':
        gene_cols[g] = 143
    else:
        gene_cols[g] = hdr.index(g)
hk_cols  = [hdr.index(h) for h in HK_NAMES]
idx_site = hdr.index('Site')
idx_data = hdr.index('Data.Chen')
idx_y    = hdr.index('Chen')

X_raw_list, HK_list, y_list, ids = [], [], [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[idx_site] != 'UCSF' or row[idx_data] != 1:
        continue
    gene_vals = [row[gene_cols[g]] for g in GENE_NAMES]
    hk_vals   = [row[c] for c in hk_cols]
    if any(v is None for v in gene_vals + hk_vals) or row[idx_y] is None:
        continue
    X_raw_list.append(gene_vals)
    HK_list.append(hk_vals)
    y_list.append(row[idx_y])
    ids.append(row[0])

X_raw = np.array(X_raw_list, dtype=float)
HK    = np.array(HK_list, dtype=float)
y     = np.array(y_list, dtype=float)
sample_ids = np.array(ids)
X_norm = hk_normalize(X_raw, HK)

print(f"Training cohort loaded: N = {X_norm.shape[0]} samples, {X_norm.shape[1]} genes")
print(f"  y ('Chen' column; rescaled LASSO risk score): min={y.min():.4f}  max={y.max():.4f}  mean={y.mean():.4f}")
print(f"  X_norm (log2 HK-normalized): mean={X_norm.mean():.3f}  std={X_norm.std():.3f}")
print(f"  Dropped samples with missing data: {sum(1 for _ in ws.iter_rows(min_row=2)) - X_norm.shape[0]}")

out_path = os.path.join(DATA_DIR, 'training_173.npz')
np.savez(out_path,
         X_norm=X_norm, X_raw=X_raw, HK=HK, y=y,
         sample_ids=sample_ids,
         gene_names=np.array(GENE_NAMES), hk_names=np.array(HK_NAMES))
print(f"\nWrote {out_path}")
